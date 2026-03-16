from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

wb = Workbook()

header_font_white = Font(bold=True, size=11, name='Arial', color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E79')
data_font = Font(size=10, name='Arial')
blue_font = Font(size=10, name='Arial', color='0000FF')
red_font = Font(size=10, name='Arial', color='CC0000')
green_font = Font(size=10, name='Arial', color='008000')
bold_font = Font(bold=True, size=10, name='Arial')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# ── Sheet 1: Statement ──
ws = wb.active
ws.title = 'Statement'
ws.merge_cells('A1:E1')
ws['A1'] = 'FIRST NATIONAL BANK'
ws['A1'].font = Font(bold=True, size=16, name='Arial', color='1F4E79')
ws.merge_cells('A2:E2')
ws['A2'] = 'Monthly Statement - March 2026'
ws['A2'].font = Font(size=11, name='Arial', color='666666')
ws.merge_cells('A3:E3')
ws['A3'] = 'Account: ****4821  |  Checking  |  03/01/2026 - 03/31/2026'
ws['A3'].font = Font(size=9, name='Arial', color='999999')

ws['A5'] = 'Beginning Balance'
ws['B5'] = 1247.33
ws['A6'] = 'Total Deposits'
ws['B6'] = '=SUMIF(D11:D70,">0",D11:D70)'
ws['A7'] = 'Total Withdrawals'
ws['B7'] = '=SUMIF(D11:D70,"<0",D11:D70)'
ws['A8'] = 'Ending Balance'
ws['B8'] = '=B5+B6+B7'
for r in range(5, 9):
    ws[f'A{r}'].font = bold_font
    ws[f'B{r}'].font = blue_font
    ws[f'B{r}'].number_format = '$#,##0.00'

headers = ['Date', 'Description', 'Category', 'Amount', 'Balance']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=10, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14

txns = [
    (date(2026,3,1), 'DIRECT DEPOSIT - ACME CORP PAYROLL', 'Income', 2500.00),
    (date(2026,3,1), 'ZELLE TRANSFER TO ROOMMATE - RENT', 'Rent', -1400.00),
    (date(2026,3,2), 'GEICO AUTO INSURANCE', 'Insurance', -185.00),
    (date(2026,3,2), 'PLANET FITNESS MONTHLY', 'Subscriptions', -49.99),
    (date(2026,3,3), 'DOORDASH ORDER #8847291', 'Food Delivery', -34.50),
    (date(2026,3,3), 'AMAZON.COM - AIRPODS MAX', 'Shopping', -549.99),
    (date(2026,3,4), 'SHELL GAS STATION #4412', 'Gas', -52.30),
    (date(2026,3,4), 'CHIPOTLE MEXICAN GRILL', 'Dining Out', -16.42),
    (date(2026,3,5), 'WALMART SUPERCENTER #2847', 'Groceries', -127.84),
    (date(2026,3,5), 'SPOTIFY PREMIUM', 'Subscriptions', -11.99),
    (date(2026,3,6), 'UBER EATS - SUSHI KING', 'Food Delivery', -47.20),
    (date(2026,3,6), 'CHASE SAPPHIRE VISA - MIN PAYMENT', 'Debt Payment', -212.00),
    (date(2026,3,7), 'BEST BUY - PS5 GAME', 'Shopping', -69.99),
    (date(2026,3,7), "MCDONALD'S #14423", 'Dining Out', -12.87),
    (date(2026,3,8), 'THE BEER GARDEN BAR TAB', 'Bars/Nightlife', -78.50),
    (date(2026,3,8), 'UBER RIDE', 'Transportation', -24.30),
    (date(2026,3,9), 'CAPITAL ONE QUICKSILVER - MIN PMT', 'Debt Payment', -96.00),
    (date(2026,3,9), 'NETFLIX PREMIUM', 'Subscriptions', -22.99),
    (date(2026,3,10), 'TARGET - CLOTHING', 'Clothing', -134.56),
    (date(2026,3,10), 'DOORDASH ORDER #8901234', 'Food Delivery', -29.75),
    (date(2026,3,11), 'VERIZON WIRELESS', 'Phone', -95.00),
    (date(2026,3,11), 'COMCAST XFINITY INTERNET', 'Internet', -75.00),
    (date(2026,3,12), 'GRUBHUB - PIZZA HUT', 'Food Delivery', -38.90),
    (date(2026,3,12), 'BEST BUY STORE CARD - MIN PAYMENT', 'Debt Payment', -56.00),
    (date(2026,3,13), 'EXXON GAS STATION', 'Gas', -48.70),
    (date(2026,3,13), 'AMAZON.COM - GAMING KEYBOARD', 'Shopping', -159.99),
    (date(2026,3,14), 'OLIVE GARDEN', 'Dining Out', -67.34),
    (date(2026,3,14), 'AMC THEATRES - 2 TICKETS', 'Entertainment', -34.00),
    (date(2026,3,15), 'DIRECT DEPOSIT - ACME CORP PAYROLL', 'Income', 2500.00),
    (date(2026,3,15), 'SALLIE MAE STUDENT LOAN PMT', 'Debt Payment', -325.00),
    (date(2026,3,15), 'TOYOTA FINANCIAL AUTO LOAN', 'Debt Payment', -385.00),
    (date(2026,3,16), 'MARCUS PERSONAL LOAN PMT', 'Debt Payment', -150.00),
    (date(2026,3,16), 'UBER EATS - THAI BASIL', 'Food Delivery', -42.15),
    (date(2026,3,17), 'FOOT LOCKER - NIKE AIR MAX', 'Clothing', -189.99),
    (date(2026,3,17), 'STARBUCKS', 'Dining Out', -7.45),
    (date(2026,3,18), 'WALMART SUPERCENTER #2847', 'Groceries', -98.42),
    (date(2026,3,18), 'HULU + LIVE TV', 'Subscriptions', -17.99),
    (date(2026,3,19), 'APPLEBEES GRILL', 'Dining Out', -43.21),
    (date(2026,3,19), 'DOORDASH ORDER #9012345', 'Food Delivery', -31.40),
    (date(2026,3,20), 'SHELL GAS STATION #4412', 'Gas', -55.20),
    (date(2026,3,20), 'AMAZON.COM - RANDOM IMPULSE BUY', 'Shopping', -89.99),
    (date(2026,3,21), 'ROOFTOP LOUNGE - BAR TAB', 'Bars/Nightlife', -112.00),
    (date(2026,3,21), 'LYFT RIDE HOME', 'Transportation', -31.50),
    (date(2026,3,22), 'CHICK-FIL-A', 'Dining Out', -14.23),
    (date(2026,3,22), 'CRUNCHYROLL ANNUAL', 'Subscriptions', -7.99),
    (date(2026,3,23), 'COSTCO WHOLESALE', 'Groceries', -187.33),
    (date(2026,3,24), 'UBER EATS - WING STOP', 'Food Delivery', -28.60),
    (date(2026,3,24), 'TOPGOLF ENTERTAINMENT', 'Entertainment', -65.00),
    (date(2026,3,25), 'ZARA - SPRING CLOTHES', 'Clothing', -210.45),
    (date(2026,3,26), 'PANERA BREAD', 'Dining Out', -18.90),
    (date(2026,3,26), 'AMAZON.COM - ECHO DOT', 'Shopping', -49.99),
    (date(2026,3,27), 'SHELL GAS STATION #4412', 'Gas', -47.80),
    (date(2026,3,27), 'DAVE AND BUSTERS', 'Entertainment', -52.00),
    (date(2026,3,28), 'DOORDASH ORDER #9123456', 'Food Delivery', -36.85),
    (date(2026,3,28), 'BUFFALO WILD WINGS', 'Dining Out', -38.70),
    (date(2026,3,29), 'THE WHISKEY BAR', 'Bars/Nightlife', -94.00),
    (date(2026,3,29), 'UBER RIDE', 'Transportation', -19.80),
    (date(2026,3,30), 'WALMART SUPERCENTER #2847', 'Groceries', -76.45),
    (date(2026,3,31), 'AMAZON.COM - BLUETOOTH SPEAKER', 'Shopping', -79.99),
    (date(2026,3,31), 'WINGSTOP DELIVERY', 'Food Delivery', -24.50),
]

row = 11
for d, desc, cat, amt in txns:
    ws.cell(row=row, column=1, value=d).number_format = 'MM/DD/YYYY'
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=cat)
    ws.cell(row=row, column=4, value=amt).number_format = '$#,##0.00'
    if row == 11:
        ws.cell(row=row, column=5).value = f'=B5+D{row}'
    else:
        ws.cell(row=row, column=5).value = f'=E{row-1}+D{row}'
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = data_font
        ws.cell(row=row, column=col).border = thin_border
    if amt < 0:
        ws.cell(row=row, column=4).font = red_font
    else:
        ws.cell(row=row, column=4).font = green_font
    row += 1

# ── Sheet 2: Debts ──
ds = wb.create_sheet('Debts')
for i, h in enumerate(['Name', 'Balance', 'APR (%)', 'Min Payment'], 1):
    c = ds.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')

ds.column_dimensions['A'].width = 30
ds.column_dimensions['B'].width = 15
ds.column_dimensions['C'].width = 12
ds.column_dimensions['D'].width = 15

debts = [
    ('Chase Sapphire Visa', 8450.00, 24.99, 212),
    ('Capital One Quicksilver', 3200.00, 22.49, 96),
    ('Best Buy Store Card', 1875.00, 27.99, 56),
    ('Sallie Mae Student Loan', 28500.00, 6.80, 325),
    ('Toyota Auto Loan', 14200.00, 4.90, 385),
    ('Personal Loan - Marcus', 5000.00, 11.99, 150),
]
for r, (name, bal, apr, minp) in enumerate(debts, 2):
    ds.cell(row=r, column=1, value=name).font = data_font
    ds.cell(row=r, column=2, value=bal).font = blue_font
    ds.cell(row=r, column=2).number_format = '$#,##0.00'
    ds.cell(row=r, column=3, value=apr).font = blue_font
    ds.cell(row=r, column=3).number_format = '0.00'
    ds.cell(row=r, column=4, value=minp).font = blue_font
    ds.cell(row=r, column=4).number_format = '$#,##0'
    for col in range(1, 5):
        ds.cell(row=r, column=col).border = thin_border

tr = len(debts) + 2
ds.cell(row=tr, column=1, value='TOTAL').font = bold_font
ds.cell(row=tr, column=2).value = f'=SUM(B2:B{tr-1})'
ds.cell(row=tr, column=2).font = bold_font
ds.cell(row=tr, column=2).number_format = '$#,##0.00'
ds.cell(row=tr, column=4).value = f'=SUM(D2:D{tr-1})'
ds.cell(row=tr, column=4).font = bold_font
ds.cell(row=tr, column=4).number_format = '$#,##0'

# ── Sheet 3: Income ──
inc = wb.create_sheet('Income')
for i, h in enumerate(['Source', 'Monthly Amount'], 1):
    c = inc.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
inc.column_dimensions['A'].width = 30
inc.column_dimensions['B'].width = 18
inc.cell(row=2, column=1, value='ACME Corp Salary').font = data_font
inc.cell(row=2, column=2, value=5000).font = blue_font
inc.cell(row=2, column=2).number_format = '$#,##0.00'

# ── Sheet 4: Expenses ──
exp = wb.create_sheet('Expenses')
for i, h in enumerate(['Category', 'Monthly Amount'], 1):
    c = exp.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
exp.column_dimensions['A'].width = 30
exp.column_dimensions['B'].width = 18

expenses = [
    ('Rent', 1400), ('Car Insurance', 185), ('Phone Bill', 95),
    ('Internet', 75), ('Spotify + Netflix + Hulu', 45),
    ('Gym Membership', 50), ('Gas', 200), ('Groceries', 450),
    ('Dining Out', 380), ('DoorDash/UberEats', 275),
    ('Amazon Shopping', 320), ('Subscriptions (misc)', 65),
    ('Entertainment', 150), ('Clothing', 200), ('Bars/Nightlife', 180),
]
for r, (cat, amt) in enumerate(expenses, 2):
    exp.cell(row=r, column=1, value=cat).font = data_font
    exp.cell(row=r, column=2, value=amt).font = blue_font
    exp.cell(row=r, column=2).number_format = '$#,##0'
    for col in range(1, 3):
        exp.cell(row=r, column=col).border = thin_border

tr2 = len(expenses) + 2
exp.cell(row=tr2, column=1, value='TOTAL MONTHLY EXPENSES').font = bold_font
exp.cell(row=tr2, column=2).value = f'=SUM(B2:B{tr2-1})'
exp.cell(row=tr2, column=2).font = Font(bold=True, size=10, name='Arial', color='CC0000')
exp.cell(row=tr2, column=2).number_format = '$#,##0'

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sample-bank-statement.xlsx')
wb.save(out)
print(f'Created {out}')
